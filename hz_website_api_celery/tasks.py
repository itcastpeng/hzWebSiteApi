#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author:zhangcong
# Email:zc_92@sina.com

from __future__ import absolute_import, unicode_literals
from .celery import app
import requests
from bs4 import BeautifulSoup
import sys, qrcode, os, datetime
# HOST = 'http://127.0.0.1:8001'
# HOST = 'http://xmgl.zhugeyingxiao.com'
project_dir = os.path.dirname(os.getcwd())
sys.path.append(project_dir)
os.environ['DJANGO_SETTINGS_MODULE'] = 'hzWebSiteApi.settings'
import django, re, json, redis, time
django.setup()
from openpyxl import Workbook
from hurong import models
from api import models as api_models
from django.db.models.aggregates import Count
from django.db.models import Q
from publicFunc.public import send_error_msg, create_xhs_admin_response
from publicFunc.tripartite_platform_oper import QueryWhetherCallingCredentialExpired
from publicFunc.public import upload_qiniu
from publicFunc.baidu_tripartite_platform_oper import tripartite_platform_oper

# 更新小红书下拉数据
@app.task
def xiaohongshu_xiala_update_data():
    # 1、将redis中存储的下拉数据存储到数据库中
    redis_obj = redis.StrictRedis(
        host='spider_redis',
        port=1111,
        db=13,
        password="Fmsuh1J50R%T*Lq15TL#IkWb#oMp^@0OYzx5Q2CSEEs$v9dd*mnqRFByoeGZ"
    )
    redis_key = "xiaohongshu_xiala_data"
    for _ in range(redis_obj.llen(redis_key)):
        item = json.loads(redis_obj.rpop(redis_key).decode('utf8'))
        keywords = item['keywords']
        models.XiaohongshuXiaLaKeywordsChildren.objects.filter(parent__keywords=keywords).delete()
        objs = models.XiaohongshuXiaLaKeywords.objects.filter(keywords=keywords)
        if objs:
            obj = objs[0]
        else:
            obj = models.XiaohongshuXiaLaKeywords.objects.create(
                keywords=keywords,
                status=2,
                create_user_id=1
            )

        xialaci_num = 0
        query = []
        for index, i in enumerate(item['data']):
            xialaci_num += 1
            xialaci = i['text'] + " " + i['desc']
            query.append(models.XiaohongshuXiaLaKeywordsChildren(keywords=xialaci, parent=obj))
        models.XiaohongshuXiaLaKeywordsChildren.objects.bulk_create(query)
        obj.status = 2
        obj.xialaci_num = xialaci_num
        obj.update_datetime = datetime.datetime.now()
        obj.save()



    # 2、假如redis队列中没有下拉关键词，则将数据库中等待查询的下拉词存入redis队列中
    redis_key = "xiaohongshu_task_list"
    if redis_obj.llen(redis_key) == 0:
        now_date = datetime.datetime.now().strftime("%Y-%m-%d")
        q = Q(update_datetime__isnull=True) | Q(update_datetime__lt=now_date)
        objs = models.XiaohongshuXiaLaKeywords.objects.filter(q)[:200]
        for obj in objs:
            item = {
                "keywords": obj.keywords,
                "task_type": "xiaohongshu_xiala"
            }
            redis_obj.lpush(redis_key, json.dumps(item))

# 更新小红书查覆盖数据
@app.task
def xiaohongshu_fugai_update_data():
    # 1、将redis中存储的覆盖数据存储到数据库中
    print("将redis中存储的覆盖数据存储到数据库中---->", datetime.datetime.now())
    redis_obj = redis.StrictRedis(
        host='spider_redis',
        port=1111,
        db=13,
        password="Fmsuh1J50R%T*Lq15TL#IkWb#oMp^@0OYzx5Q2CSEEs$v9dd*mnqRFByoeGZ"
    )



    redis_key = "xiaohongshu_fugai_data"

    for _ in range(redis_obj.llen(redis_key)):
        data = redis_obj.rpop(redis_key)

        item = json.loads(data.decode('utf8'))
        keywords = item['keywords']
        print('keywords -->', keywords)
        page_id_list = item['page_id_list']
        total_count = item['total_count']       # 笔记次数
        # {'keywords': '隆鼻', 'total_count': 0, 'page_id_list': []}

        # 更新覆盖数据
        objs = models.XiaohongshuFugai.objects.filter(keywords=keywords)
        for obj in objs:
            flag = False
            item_data = {
                'rank': 0,
                'biji_num': total_count,
                'update_datetime': datetime.datetime.now(),
            }
            for item in page_id_list:
                obj.status = 2
                if item['id'] in obj.url:
                    flag = True
                    obj.rank = item['rank']
                    obj.is_shoulu = True
                    item_data['rank'] = item['rank']

            obj.biji_num = total_count
            obj.update_datetime = datetime.datetime.now()
            obj.save()

            now_date = datetime.datetime.now().strftime("%Y-%m-%d")
            objs = models.XiaohongshuFugaiDetail.objects.filter(keywords=obj, create_datetime__gt=now_date)
            if objs:  # 已经创建当日详情
                if flag:    # 查找到覆盖
                    objs.update(**item_data)
            else:  # 不存在
                item_data['keywords'] = obj
                models.XiaohongshuFugaiDetail.objects.create(**item_data)

        # 更新霸屏王查覆盖数据
        print("更新霸屏王查覆盖数据")
        objs = models.xhs_bpw_keywords.objects.filter(keywords=keywords)
        print('objs -->', objs)
        objs.update(update_datetime=datetime.datetime.now())
        for obj in objs:
            # obj.update_datetime = datetime.datetime.now()
            # obj.save()
            # print('obj.update_datetime -->', obj.update_datetime, obj.id)
            for item in page_id_list:
                biji_url_id = item['id']        # 抓取到的数据只有笔记的id
                rank = item['rank']             # 当前在第几名
                xhs_bpw_biji_url_objs = models.xhs_bpw_biji_url.objects.filter(
                    biji_url__contains=biji_url_id,
                    uid=obj.uid
                )

                if xhs_bpw_biji_url_objs:   # 如果存在,则表示有排名
                    xhs_bpw_biji_url_obj = xhs_bpw_biji_url_objs[0]
                    models.xhs_bpw_fugai.objects.create(
                        keywords=obj,
                        biji_url=xhs_bpw_biji_url_obj,
                        rank=rank,
                        biji_num=total_count
                    )

                else:
                    pass

        # 更新下拉笔记数据
        models.XiaohongshuXiaLaKeywords.objects.filter(keywords=keywords).update(
            biji_num=total_count,
            update_datetime=datetime.datetime.now()
        )

    # 2、假如redis队列中没有任务，则将数据库中等待查询的下拉词存入redis队列中
    redis_key = "xiaohongshu_task_list"

    # # # 小红书工具单独使用(优先查询)
    # redis_obj2 = redis.StrictRedis(
    #     host='192.168.10.64',
    #     port=6379,
    #     db=4,
    # )
    # for _ in range(redis_obj2.llen("xhs_tool_api::keyword_list::xhs")):
    #     print("添加小红书工具单独使用关键词")
    #     keywords = redis_obj2.rpop("xhs_tool_api::keyword_list::xhs").decode('utf8')
    #     item = {
    #         "keywords": keywords,
    #         # "url": obj.url,
    #         # "count": 2,  # 当前关键词存在几个任务
    #         # "select_type": obj.select_type,
    #         "task_type": "xhs_tool"
    #     }
    #     redis_obj.rpush(redis_key, json.dumps(item))


    # 霸屏王查排名
    if redis_obj.llen(redis_key) == 0:

        redis_obj2 = redis.StrictRedis(host='redis', port=6381, db=0, decode_responses=True)
        keys = redis_obj2.keys("XHS_SCREEN*")
        uid_list = []
        for key in keys:
            uid = key.replace('XHS_SCREEN_', "")
            uid_list.append(uid)

        now_date = datetime.datetime.now().strftime("%Y-%m-%d")
        q = Q(update_datetime__isnull=True) | Q(update_datetime__lt=now_date)
        print('q -->', q)
        objs = models.xhs_bpw_keywords.objects.filter(q).filter(uid__in=uid_list).order_by('?')[:2000]
        print("霸屏王查排名---->", datetime.datetime.now(), objs.count())
        for obj in objs:
            print('obj.keywords----------------> ', obj.keywords)
            item = {
                "keywords": obj.keywords,
                "count": 1,  # 当前关键词存在几个任务
                "task_type": "xiaohongshu_fugai_bpw"
            }
            redis_obj.lpush(redis_key, json.dumps(item))
    #
    if redis_obj.llen(redis_key) == 0:
        # objs = models.XiaohongshuFugai.objects.all().values('keywords').annotate(Count('id'))
        objs = models.XiaohongshuFugai.objects.filter(task_type=2).values('keywords').annotate(Count('id'))
        print("普通关键词查排名---->", datetime.datetime.now(), objs.count())
        for obj in objs:
            now_date = datetime.datetime.now().strftime("%Y-%m-%d")
            # print('obj -->', obj)
            detail_objs = models.XiaohongshuFugaiDetail.objects.filter(keywords__keywords=obj['keywords'], create_datetime__gt=now_date)[:2000]
            # 将今天未查询的任务放入redis队列中
            if not detail_objs:
                item = {
                    "keywords": obj['keywords'],
                    # "url": obj.url,
                    "count": obj['id__count'],      # 当前关键词存在几个任务
                    # "select_type": obj.select_type,
                    "task_type": "xiaohongshu_fugai"
                }
                redis_obj.lpush(redis_key, json.dumps(item))

    # 相同关键词,不同链接,白天新添加的词单独跑
    print("相同关键词,不同链接,白天新添加的词单独跑---->", datetime.datetime.now())
    if redis_obj.llen(redis_key) == 0:
        objs = models.XiaohongshuFugai.objects.filter(update_datetime__isnull=True)
        for obj in objs:
            item = {
                "keywords": obj.keywords,
                # "url": obj.url,
                "count": 2,  # 当前关键词存在几个任务
                # "select_type": obj.select_type,
                "task_type": "xiaohongshu_fugai"
            }
            redis_obj.lpush(redis_key, json.dumps(item))



# 保存下拉和覆盖的数据到报表中
@app.task
def xiaohongshu_shengcheng_baobiao():
    # 保存下拉报表
    objs = models.XiaohongshuXiaLaKeywords.objects.filter(status=2)
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="编号")
    ws.cell(row=1, column=2, value="关键词名称")
    ws.cell(row=1, column=3, value="笔记数量")
    ws.cell(row=1, column=4, value="下拉词数量")
    ws.cell(row=1, column=5, value="下拉词")
    row = 2

    for obj in objs:
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=obj.keywords)
        ws.cell(row=row, column=3, value=obj.biji_num)
        ws.cell(row=row, column=4, value=obj.xialaci_num)

        xialaci_list = [i[0] for i in obj.xiaohongshuxialakeywordschildren_set.values_list('keywords')]
        ws.cell(row=row, column=5, value="\n".join(xialaci_list))
        row += 1
    excel_path = "statics/api_hurong/xiaohongshu_xiala.xlsx"
    wb.save(os.path.abspath(excel_path))

    # 保存覆盖数据
    objs = models.XiaohongshuFugai.objects.filter(status=2)
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="编号")
    ws.cell(row=1, column=2, value="关键词名称")
    ws.cell(row=1, column=3, value="笔记数量")
    ws.cell(row=1, column=4, value="排名")
    ws.cell(row=1, column=5, value="搜索类型")
    row = 2

    for obj in objs:
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=obj.keywords)
        ws.cell(row=row, column=3, value=obj.biji_num)
        ws.cell(row=row, column=4, value=obj.rank)
        ws.cell(row=row, column=5, value=obj.get_select_type_display())

        row += 1
    excel_path = "statics/api_hurong/xiaohongshu_fugai.xlsx"
    wb.save(os.path.abspath(excel_path))


# 发送邮件，每间隔5分钟一次
@app.task
def hurong_send_email():
    # 将已经完成的任务状态改成已完成
    models.TaskList.objects.filter(percentage_progress=100, status=2).update(status=3)

    # # 开始任务
    # task_list_objs = models.TaskList.objects.exclude(percentage_progress=100).filter(is_delete=False)[0: 1]
    # if task_list_objs:
    #     # 发送标题和内容
    #     task_list_obj = task_list_objs[0]
    #     task_list_obj.status = 2
    #     task_list_obj.save()
    #
    #     send_email_title = task_list_obj.send_email_title
    #     send_email_content = task_list_obj.send_email_content
    #
    #     # 收件人列表
    #     task_info_objs = task_list_obj.taskinfo_set.filter(status=1)[0: 5]
    #     task_info_id_list = []
    #     send_email_list = []
    #     for task_info in task_info_objs:
    #         task_info_id_list.append(task_info.id)
    #         send_email_list.append(task_info.to_email)
    #
    #     while True:
    #         email_user_obj = models.EmailUserInfo.objects.all().order_by('use_number')[0]
    #         email_user_obj.use_number += 1
    #         email_user_obj.save()
    #         email_user = email_user_obj.email_user
    #         email_pwd = email_user_obj.email_pwd
    #         print(email_user, email_pwd, send_email_list)
    #         send_email_obj = SendEmail(
    #             email_user,
    #             email_pwd,
    #             send_email_list,
    #             send_email_title,
    #             send_email_content
    #         )
    #         send_email_obj.send_email()
    #         if send_email_obj.send_status:
    #             # 计算完成百分比
    #             models.TaskInfo.objects.filter(id__in=task_info_id_list).update(status=2)
    #             task_objs = models.TaskInfo.objects.filter(task_list=task_list_obj)
    #             is_send_count = task_objs.filter(status=2).count()  # 已经发送成功的总数
    #             count = task_objs.count()   # 该任务的总任务数
    #             # print(is_send_count, count, is_send_count / count)
    #             task_list_obj.percentage_progress = int(is_send_count / count * 100)
    #             task_list_obj.save()
    #             break
    #         else:
    #             email_user_obj.is_available = False
    #             email_user_obj.save()


@app.task
def debug_test():
    objs = models.XiaohongshuFugai.objects.all()
    for obj in objs:
        now_date = datetime.datetime.now().strftime("%Y-%m-%d")
        objs = models.XiaohongshuFugaiDetail.objects.filter(keywords=obj, create_datetime__gt=now_date)
        if not objs:
            print('obj -->', obj.id, obj.keywords, obj.url, obj.select_type)


# 小红书手机监控
@app.task
def xiaohongshu_phone_monitor():
    # hour_now = datetime.datetime.now().strftime("%H")  # 当前的时间
    # if 7 < int(hour_now) < 21:  # 只在 8:00 - 21:00 运行
    err_phone = []
    # expire_date = (datetime.datetime.now() - datetime.timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S")
    expire_date = (datetime.datetime.now() - datetime.timedelta(minutes=5))
    phone_objs = models.XiaohongshuPhone.objects.filter(is_debug=False)
    for phone_obj in phone_objs:

        # # 如果5分钟之内没有提交日志，说明机器异常了
        # objs = phone_obj.xiaohongshuphonelog_set.filter(create_datetime__gt=expire_date)
        if phone_obj.last_sign_in_time and phone_obj.last_sign_in_time < expire_date :
            phone_obj.status = 2
            seconds = (expire_date - phone_obj.last_sign_in_time).seconds

            name = phone_obj.name
            if seconds > 60 * 60 * 2:
                name = name + " 异常超过2小时"
            elif seconds > 60 * 60 * 1:
                name = name + " 异常超过1小时"
            else:
                name = name + " 异常%s分钟" % (int(seconds / 60))

            err_phone.append(name)
        else:
            phone_obj.status = 1
        phone_obj.save()
        # print("err_phone -->", err_phone)

    if len(err_phone) > 0:
        content = """{time} \n 小红书机器异常{phone}台，请及时处理:  \n{phone_names}""".format(
            phone_names="\n".join(err_phone),
            phone=len(err_phone),
            time=datetime.datetime.today()
        )
        send_error_msg(content)  # 发送异常




# 手机号同步
@app.task
def sync_phone_number():
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36"
    }

    # 登录

    data_list = [
        {
            "login_url" : "http://47.110.86.5:9999",
            "username": "张聪296",
            "password": "zhang_cong.123",
        },{
            "login_url" : "http://120.55.80.27:9999",
            "username": "张聪0627",
            "password": "zhang_cong.123",
        }
    ]
    for data in data_list:
        login_url = data.get('login_url') + '/index.php?g=cust&m=login&a=dologin'

        requests_obj = requests.Session()
        requests_obj.post(url=login_url, headers=headers, data=data)
        # print(ret.text)

        # 获取卡号列表
        headers["Referer"] = login_url
        phone_list_url = data.get('login_url') + "/index.php?g=cust&m=cardno_cust&a=sub"

        while True:
            ret = requests_obj.get(phone_list_url, headers=headers)
            soup = BeautifulSoup(ret.text, 'lxml')
            tbody_html = soup.find('tbody')

            for tr_html in tbody_html.find_all('tr'):
                phone_num = tr_html.find_all('td')[1].text
                expire_date = tr_html.find_all('td')[2].text
                # remark = tr_html.find_all('td')[3].text
                # print(phone_num, expire_date, remark)
                if not models.PhoneNumber.objects.filter(phone_num=phone_num):
                    models.PhoneNumber.objects.create(
                        phone_num=phone_num,
                        expire_date=expire_date,
                        # remark=remark,
                    )

            next_page_html = soup.find('a', text="下一页")
            headers["Referer"] = phone_list_url
            if next_page_html:
                phone_list_url = data.get('login_url') + next_page_html.attrs["href"]
            else:
                break


# 小红书账号注册监控,监控是否有未注册的账号
@app.task
def xiaohongshu_userprofile_register_monitor():
    hour_now = datetime.datetime.now().strftime("%H")  # 当前的时间
    # if 7 < int(hour_now) < 21:  # 只在 8:00 - 21:00 运行
    from hurong import models
    objs = models.XiaohongshuUserProfileRegister.objects.filter(is_register=False)
    if objs:
        content = """{} \n小红书有新的账号需要注册，请及时处理""".format(datetime.datetime.today())
        send_error_msg(content, 3)


# 小红书未发布笔记监控,监控有新的笔记需要发布
@app.task
def xiaohongshu_biji_monitor():
    hour_now = datetime.datetime.now().strftime("%H")  # 当前的时间
    # if 7 < int(hour_now) < 21:  # 只在 8:00 - 21:00 运行
    from hurong import models
    # user_id_id = 5 是测试账号
    # objs = models.XiaohongshuBiji.objects.exclude(status=2).exclude(user_id_id=5)
    objs = models.XiaohongshuBiji.objects.filter(status=3).exclude(user_id_id=5)
    if objs:
        content = """{} \n 小红书有新的笔记需要发布，请及时处理""".format(datetime.datetime.today())
        send_error_msg(content, 3)


# 同步小红书霸屏王关键词和链接
@app.task
def xhs_bpw_keywords_rsync():
    redis_obj = redis.StrictRedis(host='redis', port=6381, db=0, decode_responses=True)
    keys = redis_obj.keys("XHS_SCREEN*")
    num = 0
    uid_list = []
    for key in keys:
        num += 1
        uid = key.replace('XHS_SCREEN_', "")
        uid_list.append(uid)
        data = redis_obj.get(key)
        data = json.loads(data)
        links = data["links"]

        if len(links) > 0:
            keywords = data["keywords"]
            query_list = []
            for keyword in keywords:
                keyword = keyword.strip()
                if not models.xhs_bpw_keywords.objects.filter(uid=uid, keywords=keyword):
                    query_list.append(models.xhs_bpw_keywords(uid=uid, keywords=keyword))
            models.xhs_bpw_keywords.objects.bulk_create(query_list)
            query_list = []
            print('=======================================================================================================获取链接', num)
            for link in links:
                # 处理短链接
                # while_flag = 0
                # while True:
                #     while_flag += 1
                try:
                    if link.startswith("http://t.cn"):
                        ret = requests.get(link, allow_redirects=False, timeout=10)
                        link = re.findall('HREF="(.*?)"', ret.text)[0].split('?')[0]
                except Exception:
                    pass
                # if while_flag>=5:
                #     break

                if not models.xhs_bpw_biji_url.objects.filter(uid=uid, biji_url=link):
                    query_list.append(models.xhs_bpw_biji_url(uid=uid, biji_url=link))
            models.xhs_bpw_biji_url.objects.bulk_create(query_list)
    models.xhs_bpw_keywords.objects.exclude(uid__in=uid_list).delete()

# 同步小红书霸屏王关键词覆盖数据到redis中
@app.task
def xhs_bpw_keywords_fugai_rsync():
    # ============================================================================
    objs = models.xhs_bpw_biji_url.objects.filter(create_datetime__isnull=False)
    for obj in objs:
        flag = 0
        biji_url = obj.biji_url
        # print('link====================> ', biji_url)
        while True:
            flag += 1
            if 'www' not in biji_url:
                if biji_url.startswith("http://t.cn"):
                    num = 0
                    while True:
                        num += 1
                        try:
                            ret = requests.get(biji_url, allow_redirects=False)
                            biji_url = re.findall('HREF="(.*?)"', ret.text)[0].split('?')[0]
                        except Exception:
                            pass
                        if num >= 3:
                            break
                else:
                    try:
                        biji_url = biji_url.split('?')[0]
                    except Exception:
                        pass

            if flag >= 5:
                break

        obj.biji_url = biji_url
        obj.save()
    # ================================================================================

    redis_obj = redis.StrictRedis(host='redis', port=6381, db=0, decode_responses=True)

    now_date = datetime.datetime.now().strftime("%Y-%m-%d")

    objs = models.xhs_bpw_fugai.objects.select_related('keywords', 'biji_url').filter(create_datetime__gt=now_date)

    data = {}
    for obj in objs:
        uid = obj.keywords.uid
        keywords = obj.keywords.keywords
        biji_url = obj.biji_url.biji_url
        rank = obj.rank
        biji_num = obj.biji_num

        keywords_data = {
            "keywords": keywords,
            "biji_url": biji_url,
            "rank": rank,
            "biji_num": biji_num,
        }
        if data.get(uid):   # 表示已经存在
            data[uid].append(keywords_data)
        else:
            data[uid] = [keywords_data]
    for uid, keywords_data in data.items():
        key = "XHS_FUGAI_{now_date}_{uid}".format(now_date=now_date, uid=uid)
        ex_seconds = 60 * 15    # key 失效的时间是15分钟
        redis_obj.set(key, json.dumps(keywords_data), ex_seconds)
        print("key -->", key)
        # print("json.dumps(keywords_data) -->", json.dumps(keywords_data))

# 定时删除 设备日志(保留近三天)
@app.task
def delete_phone_log():
    url = 'https://xcx.bjhzkq.com/api_hurong/celery/delete_phone_log'
    requests.get(url)

# 获取 手机号短信
@app.task
def celery_get_phone_content():
    url = 'https://xcx.bjhzkq.com/api_hurong/celery/celery_get_phone_content'
    requests.get(url)

# 查询设备流量信息
@app.task
def get_traffic_information():
    # url = 'http://127.0.0.1:8009/api_hurong/celery/get_traffic_information'
    url = 'https://xcx.bjhzkq.com/api_hurong/celery/get_traffic_information'
    requests.get(url)

# 给小红书后台传递数据
@app.task
def asynchronous_transfer_data(data):
    url = 'https://xcx.bjhzkq.com/api_hurong/celery/asynchronous_transfer_data'
    requests.post(url, data=data)

# 异步上传手机抓取的评论
@app.task
def error_asynchronous_transfer_data():
    url = 'https://xcx.bjhzkq.com/api_hurong/celery/error_asynchronous_transfer_data'
    requests.post(url)


# 异步 同步 日记反链
@app.task
def asynchronous_synchronous_trans(task_id):
    q = Q()
    if task_id:
        q.add(Q(id=task_id), Q.AND)
    else:
        q.add(Q(status=1), Q.AND)

    objs = models.XiaohongshuBiji.objects.filter(q,user_id__isnull=False, biji_url__isnull=False)
    api_url = "https://www.ppxhs.com/api/v1/sync/sync-screen-article"
    for obj in objs:
        data = {
            "id": obj.id,
            "link": obj.biji_url,
            "pubTime": obj.release_time,
            "from_blogger": '1',
            "platform": obj.user_id.platform,
        }
        ret = requests.post(url=api_url, data=data)
        models.AskLittleRedBook.objects.create(  # 更新日志
            request_type=2,  # POST请求
            request_url=api_url,
            get_request_parameter='',
            post_request_parameter=data,
            response_data=json.dumps(ret.json()),
            status=1
        )

# 手机号 未使用的低于200 告警
@app.task
def unused_cell_phone_number_below_alarms():
    count = models.PhoneNumber.objects.filter(status=1).count()
    if int(count) < 200:
        content = '{}\n 手机号未使用的剩余{}个'.format(
            datetime.datetime.today(),
            count
        )
        send_error_msg(content, 4)

# celery任务过多告警
@app.task
def celery_task_toomuch_alarm():
    redis_obj = redis.StrictRedis(host='redis', port=6379, db=15, decode_responses=True)
    celery_task_num = redis_obj.llen('celery')
    if celery_task_num:
        celery_task_num = int(celery_task_num)
        if celery_task_num >= 1000:
            send_error_msg('celery任务大于1000条 请及时处理 {}'.format(celery_task_num), 5)

# 查询手机号平台 判断设备 手机号是自己的还是客户的
@app.task
def determine_phone_number_ownership():
    url = 'https://xcx.bjhzkq.com/api_hurong/celery/determine_phone_number_ownership'
    requests.post(url)

# 定时刷新转接 时间是否过期
@app.task
def time_refresh_switch():
    url = 'https://xcx.bjhzkq.com/api/celery/time_refresh_switch'
    requests.get(url)

# 初始化模板 生成小程序二维码
@app.task
def get_xcx_qrcode(template_id, user_id, token):
    request_url = 'pages/index/tarBar01?token={}&user_id={}&template_id={}'.format(
        token, user_id, template_id
    )
    credential_expired_data = QueryWhetherCallingCredentialExpired('wx700c48cb72073e61', 2)  # 判断调用凭证是否过期 (操作 GZH/XCX 前调用该函数)
    authorizer_access_token = credential_expired_data.get('authorizer_access_token')

    url = 'https://api.weixin.qq.com/cgi-bin/wxaapp/createwxaqrcode?access_token={}'.format(authorizer_access_token)
    data = {
        'path': request_url,
        'width': 430
    }
    ret = requests.post(url, data=json.dumps(data))

    img_path = str(int(time.time())) + '.png'
    with open(img_path, 'wb') as f:
        f.write(ret.content)
    path = upload_qiniu(img_path, 800)
    api_models.Template.objects.filter(id=template_id).update(xcx_qrcode=path)

# 初始化模板 生成小程序二维码
@app.task
def get_gzh_qrcode(template_id, qrcode_path):
    img = qrcode.make(qrcode_path)
    time_name = str(int(time.time())) + '.png'
    with open('{}'.format(time_name), 'wb') as f:
        img.save(f)
    path = upload_qiniu(time_name, 800)
    api_models.Template.objects.filter(id=template_id).update(qrcode=path)

# 初始化模板 生成百度小程序二维码
@app.task
def get_baidu_xcx_qicode(template_id, user_id, token):
    objs = api_models.BaiduSmallProgramManagement.objects.filter(appid='14794638')
    tripartite_platform = tripartite_platform_oper()
    response = tripartite_platform.get_template_list(1, 10) # 获取模板列表
    response_data = response.data.get('list')[0]
    data = {
        'appid': objs[0].appid,
        'token': objs[0].access_token,
        'version': response_data.get('user_version'),
        'template_id': response_data.get('template_id'),
        'id': template_id,
        'user_id': user_id,
        'user_token': token,
    }
    print('data------------> ', data)
    tripartite_platform.upload_small_program_code(data)
    time.sleep(3)
    response_data = tripartite_platform.gets_list_small_packages(objs[0].access_token)
    package_id = response_data.data[0].get('package_id')
    path = tripartite_platform.get_qr_code(package_id, 200, objs[0].access_token)
    print('path-----------------------> ', path)
    api_models.Template.objects.filter(id=template_id).update(baidu_xcx_qrcode=path)

# 定时删除收录
@app.task
def timed_deletion():
    models.XiaohongshuFugai.objects.filter(task_type=2).delete()



@app.task
def generate_business_card_poster(card_id): # 生成海报
    url = 'https://xcx.bjhzkq.com/api/celery/generate_business_card_poster?card_id={}'.format(card_id)
    requests.get(url)

@app.task
def send_wechat_msg(): # 发送消息通知
    url = 'https://xcx.bjhzkq.com/api/celery/send_wechat_msg'
    requests.get(url)



@app.task
def zhangcong_test():
    objs = models.xhs_bpw_biji_url.objects.filter(biji_url__contains="xhsurl.com")
    for obj in objs:
        biji_url = obj.biji_url.strip().replace('，复制本条信息，打开【小红书】App查看精彩内容！', '').rstrip("，")
        print(biji_url)
        ret = requests.head(biji_url, allow_redirects=False)
        print(ret.headers)
        print(ret.headers['Location'])
        if "www.xiaohongshu.com" in ret.headers['Location']:
            obj.biji_url = ret.headers['Location']
            obj.save()





