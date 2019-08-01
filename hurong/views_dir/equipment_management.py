from hurong import models
from publicFunc import Response
from publicFunc import account
from django.http import JsonResponse
from publicFunc.condition_com import conditionCom
from hurong.forms.equipment_management import AddForm, SelectForm, UpdateForm
from hz_website_api_celery.tasks import get_traffic_information
from django.db.models import Q
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
import requests, datetime, json, os


# 设备管理 (手机 流量 操作 查询充值话费)


@account.is_token(models.UserProfile)
def equipment_management(request):
    response = Response.ResponseObj()
    if request.method == "GET":
        forms_obj = SelectForm(request.GET)
        if forms_obj.is_valid():
            user_id = request.GET.get('user_id')
            current_page = forms_obj.cleaned_data['current_page']
            length = forms_obj.cleaned_data['length']
            order = request.GET.get('order', '-create_datetime')
            field_dict = {
                'id': '',
                'cardbaldata': '__contains',
                'select_number': '__contains',
                'cardnumber': '__contains',
                'phone__name': '__contains',
                'create_datetime': '',
            }

            q = conditionCom(request, field_dict)
            status = request.GET.get('status')
            if status:
                if status in [1, '1']:
                    q.add(Q(cardstatus='已激活'), Q.AND)
                else:
                    q.add(Q(cardstatus='已停用'), Q.AND)

            # print('q -->', q)
            objs = models.MobileTrafficInformation.objects.filter(q).order_by(order)
            # print(objs)
            count = objs.count()

            if length != 0:
                start_line = (current_page - 1) * length
                stop_line = start_line + length
                objs = objs[start_line: stop_line]

            ret_data = []
            for obj in objs:
                cardstartdate = obj.cardstartdate
                if obj.cardstartdate:
                    cardstartdate = obj.cardstartdate.strftime('%Y-%m-%d %H:%M:%S')

                cardenddate = obj.cardenddate
                if obj.cardenddate:
                    cardenddate = obj.cardenddate.strftime('%Y-%m-%d %H:%M:%S')

                name = ''
                if obj.phone:
                    name = obj.phone.name

                phone_id = ''
                phone_number = ''
                phone_name = ''
                if obj.phone:
                    phone_number = obj.phone.phone_num
                    phone_name = obj.phone.name
                    phone_id = obj.phone_id

                ret_data.append({
                    'id': obj.id,
                    'name': name,
                    'phone_id': phone_id,
                    'phone_name': phone_name,
                    'phone_number': phone_number,
                    'cardimsi': obj.cardimsi,
                    'cardstatus': obj.cardstatus,
                    'cardtype': obj.cardtype,
                    'cardusedata': obj.cardusedata,
                    'cardno': obj.cardno,
                    'cardbaldata': obj.cardbaldata,
                    'select_number': obj.select_number,
                    'cardnumber': obj.cardnumber,
                    'cardstartdate': cardstartdate,
                    'cardenddate': cardenddate,
                    'errmsg': obj.errmsg,
                    'create_datetime': obj.create_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                })
            #  查询成功 返回200 状态码
            response.code = 200
            response.msg = '查询成功'
            response.data = {
                'ret_data': ret_data,
                'data_count': count,
                'status_choices':[{'id':1, 'name':'已激活'}, {'id':2, 'name':'已停用'}],
            }
            response.note = {
                'id': 'ID',
                'name': '设备名称',
                'cardimsi': 'ISMI号',
                'cardstatus': '用户状态',
                'cardtype': '套餐类型',
                'cardusedata': '已用流量',
                'cardno': '卡编号',
                'cardbaldata': '剩余流量',
                'select_number': '查询号码',
                'cardnumber': '卡号',
                'cardstartdate': '卡开户时间',
                'cardenddate': '卡到期时间',
                'errmsg': '错误日志',
                'create_datetime': '创建时间',
            }
        else:
            response.code = 301
            response.msg = "请求异常"
            response.data = json.loads(forms_obj.errors.as_json())

    return JsonResponse(response.__dict__)


@account.is_token(models.UserProfile)
def equipment_management_oper(request, oper_type, o_id):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    # print('request.POST -->', request.POST)
    if request.method == "POST":

        form_data = {
            'o_id': o_id,
            'select_number': request.POST.get('select_number'),
        }

        # 添加
        if oper_type == "add":
            forms_obj = AddForm(form_data)
            if forms_obj.is_valid():
                select_number = forms_obj.cleaned_data.get('select_number')
                for i in select_number:
                    models.MobileTrafficInformation.objects.create(
                        select_number=i
                    )
                get_traffic_information.delay()
                response.code = 200
                response.msg = '创建成功, 共创建{}条数据'.format(len(select_number))

            else:
                response.code = 301
                response.msg = json.loads(forms_obj.errors.as_json())

        elif oper_type == 'update':
            form_obj = UpdateForm(form_data)
            if form_obj.is_valid():
                select_number = form_obj.cleaned_data.get('select_number')
                models.MobilePhoneRechargeInformation.objects.filter(equipment_id=o_id).delete()
                objs = models.MobileTrafficInformation.objects.filter(id=o_id)
                objs.update(**{
                    'select_number':select_number,
                    'cardbaldata': '',
                    'cardenddate': None,
                    'cardimsi': '',
                    'cardno':'',
                    'cardnumber':'',
                    'cardstatus':'',
                    'cardstartdate':None,
                    'cardtype':'',
                    'cardusedata':'',
                })

                response.code = 200
                response.msg = '修改成功'

            else:
                response.code = 301
                response.msg = json.loads(form_obj.errors.as_json())

        # 删除
        elif oper_type == 'delete':
            models.MobilePhoneRechargeInformation.objects.filter(equipment_id=o_id).delete()
            models.MobileTrafficInformation.objects.filter(id=o_id).delete()
            response.code = 200
            response.msg = '删除成功'

        else:

            response.code = 402
            response.msg = "请求异常"

    else:

        # 查询设备充值信息
        if oper_type == 'get_recharge_information':
            forms_obj = SelectForm(request.GET)
            if forms_obj.is_valid():
                user_id = request.GET.get('user_id')
                current_page = forms_obj.cleaned_data['current_page']
                length = forms_obj.cleaned_data['length']
                order = request.GET.get('order', '-create_datetime')
                field_dict = {
                    'id': '',
                    'equipment_id': '',
                    'equipment_package': '',
                    'create_datetime': '__contains',
                }

                q = conditionCom(request, field_dict)

                objs = models.MobilePhoneRechargeInformation.objects.filter(q).order_by(order)
                count = objs.count()
                if length != 0:
                    start_line = (current_page - 1) * length
                    stop_line = start_line + length
                    objs = objs[start_line: stop_line]

                ret_data = []
                for obj in objs:

                    ret_data.append({
                        'equipment_id': obj.equipment_id,
                        'prepaid_phone_time': obj.equipment_package,
                        'equipment_package': obj.prepaid_phone_time.strftime('%Y-%m-%d %H:%M:%S'),
                        'create_datetime': obj.create_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                    })
                response.code = 200
                response.msg = '查询成功'
                response.data = {
                    'ret_data': ret_data,
                    'count': count
                }
                response.note = {
                    'equipment_id': '设备ID',
                    'prepaid_phone_time': '充值时间',
                    'equipment_package': '设备套餐',
                    'create_datetime': '创建时间',
                }

            else:
                response.code = 301
                response.msg = "请求异常"
                response.data = json.loads(forms_obj.errors.as_json())

        # 下载流量信息报表
        elif oper_type == 'download_report':
            all = request.GET.get('all')
            start_time = request.GET.get('start_time')
            stop_time = request.GET.get('stop_time')

            center = Alignment(horizontal='center', vertical='center')
            ft1 = Font(name='宋体', size=22)
            ft2 = Font(name='宋体', size=10)
            now_date = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
            wb = Workbook()
            ws = wb.active

            title = ws.cell(row=1, column=1, value="设备流量信息")
            title.font = ft1
            title.alignment = center

            ws.cell(row=2, column=6, value="查询时间:").alignment = center
            data = ['卡号', '设备名称', '用户状态', '流量信息', '套餐类型', '卡编号', '时间']
            for i in data:
                index = data.index(i) + 1
                key = ws.cell(row=4, column=index, value=i)
                key.alignment = center
                key.font = ft2


            # 合并单元格        开始行      结束行       用哪列          占用哪列
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)
            for i in range(1, 8):
                ws.merge_cells(start_row=2, end_row=3, start_column=i, end_column=i)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 70

            # print('设置行高')
            ws.row_dimensions[1].height = 50
            ws.row_dimensions[4].height = 30


            # print('文本居中')
            # ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


            ws.cell(row=2, column=7, value="{}".format(now_date)).alignment = center
            row = 5
            q = Q()
            if not all:
                q.add(Q(create_datetime__gte=start_time), Q(create_datetime__lte=stop_time), Q.AND)
            objs = models.MobileTrafficInformation.objects.filter(q, select_number__isnull=False).order_by('-create_datetime')
            for obj in objs:
                ws.row_dimensions[row].height = 30
                phone_name = ''
                if obj.phone:
                    phone_name = obj.phone.name

                traffic_information = """剩余流量:{}\n已用流量:{}""".format(obj.cardbaldata, obj.cardusedata)

                card_number = """卡编号:{}\nSIMI号{}""".format(obj.cardno, obj.cardimsi)

                time_info = """卡开户时间:{}\n卡到期时间{}""".format(obj.cardstartdate, obj.cardenddate)

                one = ws.cell(row=row, column=1, value="{}".format(obj.cardnumber))
                one.font = ft2
                one.alignment = center

                two = ws.cell(row=row, column=2, value="{}".format(phone_name))
                two.font = ft2
                two.alignment = center

                there = ws.cell(row=row, column=3, value="{}".format(obj.cardstatus))
                there.font = ft2
                there.alignment = center

                four = ws.cell(row=row, column=4, value="{}".format(traffic_information))
                four.font = ft2
                four.alignment = center

                four = ws.cell(row=row, column=5, value="{}".format(obj.cardtype))
                four.font = ft2
                four.alignment = center

                five = ws.cell(row=row, column=6, value="{}".format(card_number))
                five.font = ft2
                five.alignment = center


                five = ws.cell(row=row, column=7, value="{}".format(time_info))
                five.font = ft2
                five.alignment = center

                row += 1

            path = os.path.join('statics', 'imgs' , '1.xlsx')
            wb.save(path)
            response.code = 200
            response.msg = '导出成功'
            response.data = 'https://xcx.bjhzkq.com/' + path

        # 下载充值记录报表
        elif oper_type == 'download_recharge_record_report':
            id=request.GET.get('id')
            center = Alignment(horizontal='center', vertical='center')
            ft1 = Font(name='宋体', size=22)
            ft2 = Font(name='宋体', size=10)
            wb = Workbook()
            ws = wb.active
            title = ws.cell(row=1, column=1, value="设备充值信息")
            title.font = ft1
            title.alignment = center

            ws.cell(row=2, column=3, value="总额:").alignment = center
            data = ['设备套餐', '充值钱数', '充值时间', '设备名称', '创建时间']
            for i in data:
                index = data.index(i) + 1
                key = ws.cell(row=4, column=index, value=i)
                key.alignment = center
                key.font = ft2

            # 合并单元格        开始行      结束行       用哪列          占用哪列
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
            for i in range(1, 8):
                ws.merge_cells(start_row=2, end_row=3, start_column=i, end_column=i)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            #
            # # print('设置行高')
            ws.row_dimensions[1].height = 50
            ws.row_dimensions[4].height = 30
            #
            # # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            row = 5
            if id:
                objs = models.MobilePhoneRechargeInformation.objects.select_related(
                    'equipment'
                ).filter(equipment_id=id)

            else:
                objs = models.MobilePhoneRechargeInformation.objects.select_related(
                    'equipment'
                ).all()
            objs = objs.order_by('-prepaid_phone_time')
            money_count = 0
            for obj in objs:
                money = 0
                if '/月' in obj.equipment_package:
                    money = obj.equipment_package.split('元')[0].split('联通')[1]
                    money_count += int(money)

                phone_name = ''
                if obj.equipment.phone:
                    phone_name = obj.equipment.phone.name

                ws.row_dimensions[row].height = 20

                one = ws.cell(row=row, column=1, value="{}".format(obj.equipment_package))
                one.font = ft2
                one.alignment = center

                two = ws.cell(row=row, column=2, value="{}".format(money))
                two.font = ft2
                two.alignment = center

                there = ws.cell(row=row, column=3, value="{}".format(obj.prepaid_phone_time))
                there.font = ft2
                there.alignment = center

                four = ws.cell(row=row, column=4, value="{}".format(phone_name))
                four.font = ft2
                four.alignment = center

                five = ws.cell(row=row, column=5, value="{}".format(obj.create_datetime.strftime('%Y-%m-%d %H:%M:%S')))
                five.font = ft2
                five.alignment = center

                row += 1
            ws.cell(row=2, column=4, value="{}".format(money_count)).alignment = center

            # path = '1.xlsx'
            path = os.path.join('statics', 'imgs', '2.xlsx')
            wb.save(path)
            response.code = 200
            response.msg = '导出成功'
            response.data = 'https://xcx.bjhzkq.com/' + path

        else:

            response.code = 402
            response.msg = "请求异常"

    return JsonResponse(response.__dict__)
